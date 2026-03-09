#!/usr/bin/env python3
"""NOC Report Assistant — syncs Jira statuses into End-of-Shift Excel report."""

import os
import sys
import re
import json
import ssl
import calendar
import datetime
import argparse
from copy import copy
from pathlib import Path
from dataclasses import dataclass
from typing import List, Optional, Tuple

from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv() -> None:
        pass

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font as _OpenpyxlFont
    from openpyxl.styles.colors import Color as _OpenpyxlColor
except ImportError:
    print("[ERROR] openpyxl required: pip install openpyxl")
    sys.exit(1)

# Jira-style link color (#0052CC)
_HYPERLINK_COLOR = _OpenpyxlColor(rgb="FF0052CC")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
VERSION = "0.1.3"

DEFAULT_REPORT_PATH = "~/Downloads/NOC endshift report.xlsx"
SHEETS = ["Night-Shift-NEW", "Day-Shift-NEW"]

TICKET_START_ROW = 8
TICKET_COLUMN = 4       # D
STATUS_COLUMN = 5       # E

TICKET_REGEX = re.compile(r'([A-Z]+-\d+)')
NOTE_REGEX = re.compile(r'\(.*?\)')
SLACK_REGEX = re.compile(r'https?://[^/]*slack\.com/')
JIRA_LINK_REGEX = re.compile(r'https?://jira\.[^/]+/')

# Jira returns "Work In Progress" — we write "IN PROGRESS"
STATUS_MAP = {
    "WORK IN PROGRESS": "IN PROGRESS",
}

# Only stop at "Permalinks" — tickets inside "Things to monitor" must be synced too
STOP_MARKERS = ("Permalinks",)

# Month abbreviation ↔ number mapping for date logic
MONTH_ABBREV = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}
MONTH_NUM_TO_ABBREV = {v: k for k, v in MONTH_ABBREV.items()}


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------
@dataclass
class TicketUpdate:
    """Result of a single ticket status check."""
    row: int
    ticket_id: str
    old_value: str
    new_value: str
    changed: bool


@dataclass
class ShiftLayout:
    """Parsed section boundaries for a single sheet."""
    from_prev_row: int      # first row of "Things to Monitor from the previous shifts"
    from_prev_end: int      # last row of that section (row before TTM header)
    ttm_row: int            # first row of "Things to monitor"
    ttm_end: int            # last row of TTM (row before Permalinks)
    permalinks_row: int     # row of the "Permalinks" header


@dataclass
class RowSnapshot:
    """Captured ticket data from a single row for cross-sheet copy."""
    summary: Optional[str]
    ticket_id: Optional[str]
    ticket_hyperlink: Optional[str]
    status: Optional[str]
    slack_text: Optional[str]
    slack_hyperlink: Optional[str]
    col_d_raw_value: Optional[str]


# ---------------------------------------------------------------------------
# Main class
# ---------------------------------------------------------------------------
class NOCReportAssistant:
    """Reads an End-of-Shift Excel report, fetches Jira statuses, and updates cells."""

    def __init__(
        self,
        jira_url: str,
        jira_token: str,
        dry_run: bool = False,
        verbose: bool = False,
    ) -> None:
        self.jira_url = jira_url.rstrip("/")
        self.jira_token = jira_token
        self.dry_run = dry_run
        self.verbose = verbose

    # -- public: sync statuses -----------------------------------------------

    def run(self, file_path: Path, sheet_name: str) -> List[TicketUpdate]:
        """Open Excel, walk ticket rows, update statuses, save."""
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
        updates: List[TicketUpdate] = []

        for row_number in range(TICKET_START_ROW, worksheet.max_row + 1):
            # Stop when we hit a section marker
            cell_a_value = str(worksheet.cell(row=row_number, column=1).value or "")
            if row_number > TICKET_START_ROW + 1 and any(
                marker in cell_a_value for marker in STOP_MARKERS
            ):
                break

            # Extract ticket ID from column D (plain text or =HYPERLINK formula)
            ticket_id = self._extract_ticket_id(
                worksheet.cell(row=row_number, column=TICKET_COLUMN)
            )
            if not ticket_id:
                continue

            # Fetch status from Jira
            jira_status, jira_assignee = self._fetch_jira_status(ticket_id)
            if jira_status is None:
                continue

            # Build new value for column E
            old_value = str(
                worksheet.cell(row=row_number, column=STATUS_COLUMN).value or ""
            )
            new_value = self._build_status_string(jira_status, jira_assignee, old_value)
            changed = old_value.strip() != new_value.strip()

            if not self.dry_run:
                worksheet.cell(row=row_number, column=STATUS_COLUMN).value = new_value

            updates.append(
                TicketUpdate(row_number, ticket_id, old_value, new_value, changed)
            )

        if not self.dry_run:
            workbook.save(file_path)

        return updates

    # -- public: add row -----------------------------------------------------

    def add_row(
        self,
        file_path: Path,
        sheet_name: str,
        jira_link: str,
        slack_link: str,
    ) -> int:
        """Add a ticket to the 'Things to monitor' section."""
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]

        # Find "Things to monitor" row (not "from the previous shifts")
        ttm_row: Optional[int] = None
        for row_number in range(TICKET_START_ROW + 2, worksheet.max_row + 1):
            cell_a_value = str(worksheet.cell(row=row_number, column=1).value or "")
            if ("Things to monitor" in cell_a_value
                    and "from" not in cell_a_value.lower()):
                ttm_row = row_number
                break

        if ttm_row is None:
            raise RuntimeError("Could not find 'Things to monitor' section")

        # Find "Permalinks" row (next section after "Things to monitor")
        permalinks_row: Optional[int] = None
        for row_number in range(ttm_row + 1, worksheet.max_row + 1):
            cell_a_value = str(worksheet.cell(row=row_number, column=1).value or "")
            if "Permalinks" in cell_a_value:
                permalinks_row = row_number
                break

        if permalinks_row is None:
            raise RuntimeError("Could not find 'Permalinks' section")

        # Reference row for formatting: last ticket row above "Things to monitor"
        reference_row = ttm_row - 1

        # Determine target row
        ttm_has_data = worksheet.cell(row=ttm_row, column=TICKET_COLUMN).value
        if not ttm_has_data:
            # First ticket — write directly into "Things to monitor" row
            target_row = ttm_row
        else:
            # Insert new row just before Permalinks
            worksheet.insert_rows(permalinks_row)
            target_row = permalinks_row
            new_permalinks_row = permalinks_row + 1

            # insert_rows may duplicate merges onto the new row — remove them
            for merge_range in list(worksheet.merged_cells.ranges):
                if (merge_range.min_row == target_row
                        and merge_range.max_row == target_row
                        and merge_range.max_col > 2):
                    worksheet.merged_cells.ranges.remove(merge_range)

            # Expand "Things to monitor" A:B merge to cover all ticket rows
            for merge_range in list(worksheet.merged_cells.ranges):
                if (merge_range.min_row == ttm_row
                        and merge_range.min_col == 1
                        and merge_range.max_col == 2):
                    worksheet.merged_cells.ranges.remove(merge_range)
                    break
            worksheet.merge_cells(
                start_row=ttm_row, start_column=1,
                end_row=target_row, end_column=2,
            )

            # Fix Permalinks full-width merge (A:F) at its new position
            for merge_range in list(worksheet.merged_cells.ranges):
                if (merge_range.min_row == new_permalinks_row
                        and merge_range.max_row == new_permalinks_row):
                    worksheet.merged_cells.ranges.remove(merge_range)
            worksheet.merge_cells(
                start_row=new_permalinks_row, start_column=1,
                end_row=new_permalinks_row, end_column=6,
            )

            _remove_hyperlinks_in_range(
                worksheet, new_permalinks_row, 2, new_permalinks_row, 6,
            )
            _ensure_permalink_merges(worksheet, new_permalinks_row)

        # Extract ticket ID from link
        ticket_id_match = TICKET_REGEX.search(jira_link)
        if not ticket_id_match:
            raise ValueError(f"Could not extract ticket ID from: {jira_link}")
        ticket_id = ticket_id_match.group(1)

        # Fetch full ticket data from Jira
        summary, jira_status, jira_assignee = self._fetch_jira_full(ticket_id)

        # C: summary
        worksheet.cell(row=target_row, column=3).value = summary
        _copy_cell_style(
            worksheet.cell(row=reference_row, column=3),
            worksheet.cell(row=target_row, column=3),
        )

        # D: ticket ID with hyperlink (native Hyperlink, not =HYPERLINK formula)
        cell_d = worksheet.cell(row=target_row, column=4)
        cell_d.value = ticket_id
        cell_d.hyperlink = jira_link
        _copy_cell_style(
            worksheet.cell(row=reference_row, column=4), cell_d,
        )
        _apply_hyperlink_font(cell_d)

        # E: STATUS Assignee
        status_string = self._build_status_string(jira_status, jira_assignee, "")
        worksheet.cell(row=target_row, column=5).value = status_string
        _copy_cell_style(
            worksheet.cell(row=reference_row, column=5),
            worksheet.cell(row=target_row, column=5),
        )

        # F: slack_link with hyperlink (native Hyperlink)
        cell_f = worksheet.cell(row=target_row, column=6)
        cell_f.value = "slack_link"
        cell_f.hyperlink = slack_link
        _copy_cell_style(
            worksheet.cell(row=reference_row, column=6), cell_f,
        )
        _apply_hyperlink_font(cell_f)

        # Columns G+ : copy fill from reference row so inserted rows don't
        # appear as black bars in Numbers/Excel (insert_rows leaves them empty).
        for extra_col in range(7, worksheet.max_column + 1):
            _copy_cell_style(
                worksheet.cell(row=reference_row, column=extra_col),
                worksheet.cell(row=target_row, column=extra_col),
            )

        workbook.save(file_path)
        return target_row

    # -- public: start shift -------------------------------------------------

    def start_shift(
        self,
        file_path: Path,
        target_sheet_name: str,
    ) -> dict:
        """Start-of-shift workflow:

        1. Update date on target sheet.
        2. Copy all tickets from source (opposite) sheet into target's
           'Things to Monitor from the previous shifts' section.
        3. Reset target's 'Things to monitor' to one empty row.
        4. Save and run sync to refresh Jira statuses.
        """
        source_sheet_name = self._opposite_sheet(target_sheet_name)

        workbook = load_workbook(file_path)
        source_ws = workbook[source_sheet_name]
        target_ws = workbook[target_sheet_name]

        # Step 1: Date update
        new_day, new_month = self._update_date(
            target_ws, source_ws, target_sheet_name,
        )

        # Step 2: Collect all tickets from source (both sections)
        source_layout = self._scan_layout(source_ws)
        source_rows = self._collect_source_rows(source_ws, source_layout)

        # Step 3: Restructure target "from previous shifts"
        target_layout = self._scan_layout(target_ws)
        current_count = target_layout.from_prev_end - target_layout.from_prev_row + 1
        source_count = max(len(source_rows), 1)  # keep at least the header row
        delta = source_count - current_count

        if delta > 0:
            # Need more rows — insert BEFORE "Things to monitor" header
            target_ws.insert_rows(target_layout.ttm_row, delta)
            # Post-insert: remove duplicate merges on newly inserted rows
            for merge_range in list(target_ws.merged_cells.ranges):
                new_start = target_layout.ttm_row
                new_end = target_layout.ttm_row + delta - 1
                if (new_start <= merge_range.min_row <= new_end
                        and merge_range.min_row == merge_range.max_row
                        and merge_range.max_col > 2):
                    target_ws.merged_cells.ranges.remove(merge_range)
        elif delta < 0:
            # Need fewer rows — delete from bottom of "from previous" section
            delete_start = target_layout.from_prev_row + source_count
            target_ws.delete_rows(delete_start, abs(delta))

        # Rescan after structural changes
        target_layout = self._scan_layout(target_ws)

        # Rebuild "from previous shifts" A:B merge
        _rebuild_section_merge(
            target_ws,
            target_layout.from_prev_row,
            target_layout.from_prev_row + source_count - 1,
        )

        # Write source data into target "from previous" rows
        reference_row = target_layout.from_prev_row
        for idx, snap in enumerate(source_rows):
            target_row = target_layout.from_prev_row + idx
            _write_ticket_row(target_ws, target_row, snap, reference_row)

        # If source was empty, clear the single remaining header row's C-F
        if not source_rows:
            for col in [3, 4, 5, 6]:
                target_ws.cell(
                    row=target_layout.from_prev_row, column=col,
                ).value = None

        # Step 4: Reset "Things to monitor" to one empty row
        target_layout = self._scan_layout(target_ws)
        extra_ttm_rows = target_layout.ttm_end - target_layout.ttm_row
        if extra_ttm_rows > 0:
            target_ws.delete_rows(target_layout.ttm_row + 1, extra_ttm_rows)
            target_layout = self._scan_layout(target_ws)

        # Clear C-F on TTM header row
        for col in [3, 4, 5, 6]:
            cell = target_ws.cell(row=target_layout.ttm_row, column=col)
            cell.value = None
            cell.hyperlink = None
        _remove_hyperlinks_in_range(
            target_ws, target_layout.ttm_row, 3, target_layout.ttm_row, 6,
        )

        # Rebuild TTM A:B merge to exactly 1 row
        _rebuild_section_merge(
            target_ws, target_layout.ttm_row, target_layout.ttm_row,
        )

        # Step 5: Repair Permalinks
        target_layout = self._scan_layout(target_ws)
        for merge_range in list(target_ws.merged_cells.ranges):
            if (merge_range.min_row == target_layout.permalinks_row
                    and merge_range.max_row == target_layout.permalinks_row):
                target_ws.merged_cells.ranges.remove(merge_range)
        target_ws.merge_cells(
            start_row=target_layout.permalinks_row, start_column=1,
            end_row=target_layout.permalinks_row, end_column=6,
        )
        _remove_hyperlinks_in_range(
            target_ws, target_layout.permalinks_row, 2,
            target_layout.permalinks_row, 6,
        )
        _ensure_permalink_merges(target_ws, target_layout.permalinks_row)

        # Step 6: Save + Sync
        if not self.dry_run:
            workbook.save(file_path)
            sync_updates = self.run(file_path, target_sheet_name)
        else:
            sync_updates = []

        return {
            "tickets_copied": len(source_rows),
            "date_day": new_day,
            "date_month": new_month,
            "sync_updates": sync_updates,
        }

    # -- internal: layout scanning -------------------------------------------

    @staticmethod
    def _scan_layout(worksheet) -> ShiftLayout:
        """Scan sheet and return section row boundaries."""
        from_prev_row: Optional[int] = None
        ttm_row: Optional[int] = None
        permalinks_row: Optional[int] = None

        for row_number in range(TICKET_START_ROW, worksheet.max_row + 1):
            cell_a_value = str(worksheet.cell(row=row_number, column=1).value or "")
            if not cell_a_value:
                continue

            cell_a_lower = cell_a_value.lower()
            if "from the previous shifts" in cell_a_lower and from_prev_row is None:
                from_prev_row = row_number
            elif ("things to monitor" in cell_a_lower
                  and "from" not in cell_a_lower
                  and ttm_row is None):
                ttm_row = row_number
            elif "permalinks" in cell_a_lower and permalinks_row is None:
                permalinks_row = row_number
                break

        if from_prev_row is None:
            raise RuntimeError(
                "Could not find 'Things to Monitor from the previous shifts' section"
            )
        if ttm_row is None:
            raise RuntimeError("Could not find 'Things to monitor' section")
        if permalinks_row is None:
            raise RuntimeError("Could not find 'Permalinks' section")

        return ShiftLayout(
            from_prev_row=from_prev_row,
            from_prev_end=ttm_row - 1,
            ttm_row=ttm_row,
            ttm_end=permalinks_row - 1,
            permalinks_row=permalinks_row,
        )

    @staticmethod
    def _collect_source_rows(
        worksheet, layout: ShiftLayout,
    ) -> List[RowSnapshot]:
        """Collect all ticket rows from both sections of a sheet."""
        rows: List[RowSnapshot] = []
        for row_number in range(layout.from_prev_row, layout.ttm_end + 1):
            cell_d = worksheet.cell(row=row_number, column=TICKET_COLUMN)
            raw_d = str(cell_d.value or "").strip()
            if not raw_d:
                continue

            cell_c = worksheet.cell(row=row_number, column=3)
            cell_e = worksheet.cell(row=row_number, column=STATUS_COLUMN)
            cell_f = worksheet.cell(row=row_number, column=6)

            d_hl = cell_d.hyperlink
            d_hyperlink = d_hl.target if d_hl and d_hl.target else None

            f_hl = cell_f.hyperlink
            f_hyperlink = f_hl.target if f_hl and f_hl.target else None

            rows.append(RowSnapshot(
                summary=cell_c.value,
                ticket_id=raw_d,
                ticket_hyperlink=d_hyperlink,
                status=str(cell_e.value or ""),
                slack_text=str(cell_f.value or ""),
                slack_hyperlink=f_hyperlink,
                col_d_raw_value=cell_d.value,
            ))
        return rows

    @staticmethod
    def _opposite_sheet(sheet_name: str) -> str:
        """Return the opposing sheet name."""
        if sheet_name == "Night-Shift-NEW":
            return "Day-Shift-NEW"
        return "Night-Shift-NEW"

    @staticmethod
    def _update_date(
        target_ws, source_ws, target_sheet_name: str,
    ) -> Tuple[int, str]:
        """Update A1 (day) and A2 (month) on the target sheet.

        Returns (new_day, new_month_str) for display.
        """
        source_day = int(source_ws.cell(row=1, column=1).value or 0)
        source_month_str = str(source_ws.cell(row=2, column=1).value or "")

        if target_sheet_name == "Night-Shift-NEW":
            new_day = source_day + 1
        else:
            new_day = source_day

        new_month_str = source_month_str

        # Handle month boundary (only when Night increments the day)
        if target_sheet_name == "Night-Shift-NEW":
            month_num = MONTH_ABBREV.get(source_month_str)
            if month_num:
                current_year = datetime.date.today().year
                _, days_in_month = calendar.monthrange(current_year, month_num)
                if new_day > days_in_month:
                    new_day = 1
                    next_month = month_num + 1 if month_num < 12 else 1
                    new_month_str = MONTH_NUM_TO_ABBREV.get(
                        next_month, source_month_str,
                    )

        target_ws.cell(row=1, column=1).value = new_day
        target_ws.cell(row=2, column=1).value = new_month_str
        return new_day, new_month_str

    # -- internal: Jira API --------------------------------------------------

    def _fetch_jira_status(self, ticket_id: str) -> Tuple[Optional[str], Optional[str]]:
        """GET /rest/api/2/issue/{ID}?fields=status,assignee — returns (status, assignee)."""
        url = f"{self.jira_url}/rest/api/2/issue/{ticket_id}?fields=status,assignee"
        data = self._jira_get(url)
        if data is None:
            return None, None

        jira_status = data["fields"]["status"]["name"]
        assignee_object = data["fields"].get("assignee")
        jira_assignee = assignee_object["displayName"] if assignee_object else "Unassigned"
        return jira_status, jira_assignee

    def _fetch_jira_full(self, ticket_id: str) -> Tuple[str, str, str]:
        """GET Jira — returns (summary, status, assignee). Raises on failure."""
        url = f"{self.jira_url}/rest/api/2/issue/{ticket_id}?fields=status,assignee,summary"
        data = self._jira_get(url)
        if data is None:
            raise RuntimeError(f"Failed to fetch Jira issue: {ticket_id}")

        summary = data["fields"].get("summary", "")
        jira_status = data["fields"]["status"]["name"]
        assignee_object = data["fields"].get("assignee")
        jira_assignee = assignee_object["displayName"] if assignee_object else "Unassigned"
        return summary, jira_status, jira_assignee

    def _jira_get(self, url: str) -> Optional[dict]:
        """Low-level Jira GET with Bearer token and SSL verification disabled."""
        request = Request(url, headers={
            "Authorization": f"Bearer {self.jira_token}",
            "Content-Type": "application/json",
        })
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE

        try:
            with urlopen(request, context=ssl_context) as response:
                return json.loads(response.read())
        except (URLError, HTTPError, KeyError) as error:
            if self.verbose:
                print(f"  [WARN] Jira request failed ({url}): {error}", file=sys.stderr)
            return None

    # -- internal: helpers ---------------------------------------------------

    def _extract_ticket_id(self, cell) -> Optional[str]:
        """Extract ticket ID from cell value (plain text or =HYPERLINK formula)."""
        raw_value = str(cell.value or "").strip()
        match = TICKET_REGEX.search(raw_value)
        return match.group(1) if match else None

    def _build_status_string(
        self, status: str, assignee: str, old_value: str
    ) -> str:
        """Format: 'STATUS_UPPERCASE Assignee Name (optional note)'."""
        status_upper = status.upper()
        # Map verbose Jira statuses to short form
        status_upper = STATUS_MAP.get(status_upper, status_upper)
        new_value = f"{status_upper} {assignee}"
        # Preserve parenthetical notes from old value
        note_match = NOTE_REGEX.search(old_value)
        if note_match:
            new_value = f"{new_value} {note_match.group(0)}"
        return new_value


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _copy_cell_style(source_cell, target_cell) -> None:
    """Copy fill, font, border, alignment and number format from source to target."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format


def _apply_hyperlink_font(cell) -> None:
    """Override font color to Jira-blue (#0052CC) with single underline for hyperlink cells."""
    old_font = cell.font
    cell.font = _OpenpyxlFont(
        name=old_font.name,
        size=old_font.size,
        bold=old_font.bold,
        italic=old_font.italic,
        underline="single",
        color=_HYPERLINK_COLOR,
        strikethrough=old_font.strikethrough,
        vertAlign=old_font.vertAlign,
    )


def _remove_hyperlinks_in_range(
    worksheet,
    min_row: int, min_col: int,
    max_row: int, max_col: int,
) -> None:
    """Remove hyperlinks whose ref falls within the given cell range."""
    from openpyxl.utils.cell import coordinate_to_tuple

    to_keep = []
    for hyperlink in worksheet._hyperlinks:
        if hyperlink.ref:
            try:
                hl_row, hl_col = coordinate_to_tuple(hyperlink.ref)
                if (min_row <= hl_row <= max_row and min_col <= hl_col <= max_col):
                    continue
            except (ValueError, TypeError):
                pass
        to_keep.append(hyperlink)
    worksheet._hyperlinks = to_keep


def _ensure_permalink_merges(worksheet, permalinks_row: int) -> None:
    """Ensure every permalink row below Permalinks has A:B and C:F merges."""
    existing_merges: set = set()
    for merge_range in worksheet.merged_cells.ranges:
        if merge_range.min_row == merge_range.max_row:
            existing_merges.add(
                (merge_range.min_row, merge_range.min_col, merge_range.max_col)
            )

    for row_number in range(permalinks_row + 1, worksheet.max_row + 1):
        cell_a_value = worksheet.cell(row=row_number, column=1).value
        cell_c_value = worksheet.cell(row=row_number, column=3).value
        if cell_a_value is None and cell_c_value is None:
            break

        if (row_number, 1, 2) not in existing_merges:
            worksheet.merge_cells(
                start_row=row_number, start_column=1,
                end_row=row_number, end_column=2,
            )
        if (row_number, 3, 6) not in existing_merges:
            worksheet.merge_cells(
                start_row=row_number, start_column=3,
                end_row=row_number, end_column=6,
            )


def _rebuild_section_merge(
    worksheet, start_row: int, end_row: int, max_col: int = 2,
) -> None:
    """Remove existing A:B merge that covers start_row, then recreate for the new range."""
    for merge_range in list(worksheet.merged_cells.ranges):
        if (merge_range.min_col == 1
                and merge_range.max_col <= max_col
                and merge_range.min_row <= start_row <= merge_range.max_row):
            worksheet.merged_cells.ranges.remove(merge_range)
    if start_row <= end_row:
        worksheet.merge_cells(
            start_row=start_row, start_column=1,
            end_row=end_row, end_column=max_col,
        )


def _write_ticket_row(
    worksheet, target_row: int, snap: RowSnapshot, reference_row: int,
) -> None:
    """Write a RowSnapshot into a target row with proper styling and hyperlinks."""
    # C: summary
    cell_c = worksheet.cell(row=target_row, column=3)
    cell_c.value = snap.summary
    _copy_cell_style(worksheet.cell(row=reference_row, column=3), cell_c)

    # D: ticket ID + hyperlink
    cell_d = worksheet.cell(row=target_row, column=4)
    cell_d.value = snap.col_d_raw_value
    _copy_cell_style(worksheet.cell(row=reference_row, column=4), cell_d)
    if snap.ticket_hyperlink:
        cell_d.hyperlink = snap.ticket_hyperlink
        _apply_hyperlink_font(cell_d)

    # E: status + assignee
    cell_e = worksheet.cell(row=target_row, column=5)
    cell_e.value = snap.status
    _copy_cell_style(worksheet.cell(row=reference_row, column=5), cell_e)

    # F: slack link
    cell_f = worksheet.cell(row=target_row, column=6)
    cell_f.value = snap.slack_text
    _copy_cell_style(worksheet.cell(row=reference_row, column=6), cell_f)
    if snap.slack_hyperlink:
        cell_f.hyperlink = snap.slack_hyperlink
        _apply_hyperlink_font(cell_f)

    # Columns G+: copy fill to prevent black bars (openpyxl pitfall #6)
    for extra_col in range(7, worksheet.max_column + 1):
        _copy_cell_style(
            worksheet.cell(row=reference_row, column=extra_col),
            worksheet.cell(row=target_row, column=extra_col),
        )


# ---------------------------------------------------------------------------
# Interactive helpers
# ---------------------------------------------------------------------------

def collect_links() -> Tuple[str, str]:
    """Collect Jira and Slack links from user in any order. Returns (jira_link, slack_link)."""
    jira_link: Optional[str] = None
    slack_link: Optional[str] = None

    while not (jira_link and slack_link):
        remaining_types: List[str] = []
        if not jira_link:
            remaining_types.append("Jira")
        if not slack_link:
            remaining_types.append("Slack")

        prompt_text = f"Paste a link ({' or '.join(remaining_types)}): "
        link = input(prompt_text).strip()

        if not link:
            continue

        if SLACK_REGEX.search(link) and not slack_link:
            slack_link = link
            print("  \u2713 Slack link detected")
        elif JIRA_LINK_REGEX.search(link) and not jira_link:
            ticket_match = TICKET_REGEX.search(link)
            if ticket_match:
                jira_link = link
                print(f"  \u2713 Jira link detected: {ticket_match.group(1)}")
            else:
                print("  \u2717 Could not extract ticket ID from Jira link")
        else:
            print("  \u2717 Unrecognized link, try again")

    return jira_link, slack_link


def select_sheet() -> str:
    """Interactive sheet selection menu. No default — user must pick a number."""
    print("Select sheet:")
    for index, sheet_name in enumerate(SHEETS, 1):
        print(f"  {index}. {sheet_name}")
    while True:
        try:
            choice = int(input("Enter number: "))
            if 1 <= choice <= len(SHEETS):
                return SHEETS[choice - 1]
        except (ValueError, EOFError):
            pass
        print(f"  Invalid choice. Enter 1-{len(SHEETS)}.")


def select_action() -> int:
    """Interactive action selection menu. No default — user must pick a number."""
    print("Select action:")
    print("  1. Start shift       \u2014 copy tickets from previous shift, update date, sync")
    print("  2. End shift (SYNC)  \u2014 sync Jira statuses for all existing tickets")
    print("  3. Add row           \u2014 add new ticket row to the report")
    while True:
        try:
            choice = int(input("Enter number: "))
            if 1 <= choice <= 3:
                return choice
        except (ValueError, EOFError):
            pass
        print("  Invalid choice. Enter 1-3.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="NOC Report Assistant — sync Jira statuses into End-of-Shift Excel report"
    )
    parser.add_argument(
        "--file", "-f",
        help=f"Path to Excel report (default: {DEFAULT_REPORT_PATH})",
    )
    parser.add_argument(
        "--dry-run", "-n",
        action="store_true",
        help="Show what would change without saving",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Show warnings and debug info",
    )
    return parser.parse_args()


def main() -> None:
    """Entry point."""
    load_dotenv()
    args = parse_args()

    # Required env vars
    jira_url = os.environ.get("JIRA_SERVER_URL", "").strip()
    jira_token = os.environ.get("JIRA_PERSONAL_ACCESS_TOKEN", "").strip()
    missing_vars: List[str] = []
    if not jira_url:
        missing_vars.append("JIRA_SERVER_URL")
    if not jira_token:
        missing_vars.append("JIRA_PERSONAL_ACCESS_TOKEN")
    if missing_vars:
        print(f"[ERROR] Missing env vars: {', '.join(missing_vars)}")
        sys.exit(1)

    # Resolve file path
    file_path = Path(
        args.file or os.environ.get("NOC_REPORT_PATH", DEFAULT_REPORT_PATH)
    )
    file_path = file_path.expanduser()
    if not file_path.exists():
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)

    # Banner
    print(f"  NOC Report Assistant v{VERSION}")
    print(f"  File: {file_path.name}")
    print()

    # Sheet selection
    sheet_name = select_sheet()
    print(f"  Sheet: {sheet_name}")
    print()

    # Action selection
    action = select_action()
    print()

    assistant = NOCReportAssistant(jira_url, jira_token, args.dry_run, args.verbose)

    if action == 1:
        # Start shift
        source_sheet = assistant._opposite_sheet(sheet_name)
        print(f"Starting shift handoff for {sheet_name}...")
        print(f"  Source: {source_sheet}")
        print(f"  Target: {sheet_name}")
        print()

        result = assistant.start_shift(file_path, sheet_name)
        tickets_copied = result["tickets_copied"]
        new_day = result["date_day"]
        new_month = result["date_month"]
        sync_updates = result["sync_updates"]
        changed_count = sum(1 for u in sync_updates if u.changed)

        print(f"  Date   : {new_month} {new_day}")
        print(f"  Copied : {tickets_copied} ticket(s) from {source_sheet}")
        if sync_updates:
            print(f"  Synced : {len(sync_updates)} tickets, {changed_count} changed")
        if args.dry_run:
            print("\n[DRY RUN] No changes saved.")
        else:
            print(f"\n  File saved: {file_path.name}")

    elif action == 2:
        # End shift (SYNC)
        updates = assistant.run(file_path, sheet_name)
        changed_count = sum(1 for update in updates if update.changed)
        print(f"\nProcessed: {len(updates)} tickets, Changed: {changed_count}")
        if args.dry_run:
            print("[DRY RUN] No changes saved.")
        for update in updates:
            if update.changed:
                print(f"  Row {update.row}: {update.ticket_id}")
                print(f"    - {update.old_value}")
                print(f"    + {update.new_value}")

    elif action == 3:
        # Add row
        jira_link, slack_link = collect_links()
        ticket_id_match = TICKET_REGEX.search(jira_link)
        if not ticket_id_match:
            print("[ERROR] Could not extract ticket ID from Jira link")
            sys.exit(1)
        ticket_id = ticket_id_match.group(1)
        print(f"\nAdding {ticket_id} to {sheet_name}...")

        if args.dry_run:
            summary, jira_status, jira_assignee = assistant._fetch_jira_full(ticket_id)
            print("  [DRY RUN] Would insert row:")
            print(f"    C: {summary}")
            print(f"    D: {ticket_id} (hyperlink)")
            print(f"    E: {assistant._build_status_string(jira_status, jira_assignee, '')}")
            print(f"    F: slack_link (hyperlink)")
        else:
            inserted_row = assistant.add_row(file_path, sheet_name, jira_link, slack_link)
            print(f"  \u2713 Row inserted at row {inserted_row}")
            print(f"  File saved: {file_path.name}")


if __name__ == "__main__":
    main()
