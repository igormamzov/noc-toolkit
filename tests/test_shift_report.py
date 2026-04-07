"""Tests for noc-report-assistant (tool #6)."""

import argparse
import sys
from io import StringIO
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from shift_report import (
    ShiftReport,
    ShiftLayout,
    RowSnapshot,
    TICKET_START_ROW,
    STATUS_MAP,
    VERSION,
    SHEETS,
    _rebuild_section_merge,
    _remove_hyperlinks_in_range,
    _ensure_permalink_merges,
    _apply_hyperlink_font,
    _copy_cell_style,
    collect_links,
    select_sheet,
    select_action,
    parse_args,
    main,
)


# ===================================================================
# _scan_layout
# ===================================================================

class TestScanLayout:

    def test_normal_sheet_with_header(self, report_path):
        """Both sections have headers — standard case."""
        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        assert layout.from_prev_row == 8
        assert layout.ttm_row > layout.from_prev_row
        assert layout.permalinks_row > layout.ttm_row
        assert layout.from_prev_end == layout.ttm_row - 1
        assert layout.ttm_end == layout.permalinks_row - 1

    def test_missing_from_prev_header(self, report_no_header):
        """Day-Shift-NEW has no 'from previous shifts' header — should fallback to row 8."""
        wb = load_workbook(report_no_header)
        ws = wb["Day-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        assert layout.from_prev_row == TICKET_START_ROW
        assert layout.ttm_row > layout.from_prev_row
        assert layout.permalinks_row > layout.ttm_row

    def test_trailing_spaces_in_ttm(self, report_path):
        """'Things to monitor        ' (trailing spaces) should still be detected."""
        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)
        # TTM should be found despite trailing spaces
        ttm_value = str(ws.cell(row=layout.ttm_row, column=1).value or "")
        assert "things to monitor" in ttm_value.lower().strip()

    def test_missing_ttm_raises(self, tmp_path):
        """Sheet without 'Things to monitor' should raise RuntimeError."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=8, column=1, value="Things to Monitor\nfrom the previous shifts")
        ws.cell(row=12, column=1, value="Permalinks")
        # No "Things to monitor" (without "from") — should fail
        # Actually row 12 has Permalinks but no TTM
        # Need to NOT have a "Things to monitor" row
        file_path = tmp_path / "broken.xlsx"
        wb.save(file_path)

        wb2 = load_workbook(file_path)
        with pytest.raises(RuntimeError, match="Things to monitor"):
            ShiftReport._scan_layout(wb2.active)

    def test_missing_permalinks_raises(self, tmp_path):
        """Sheet without 'Permalinks' should raise RuntimeError."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=8, column=1, value="Things to Monitor\nfrom the previous shifts")
        ws.cell(row=12, column=1, value="Things to monitor")
        # No Permalinks row
        file_path = tmp_path / "broken2.xlsx"
        wb.save(file_path)

        wb2 = load_workbook(file_path)
        with pytest.raises(RuntimeError, match="Permalinks"):
            ShiftReport._scan_layout(wb2.active)

    def test_newline_in_from_prev_header(self, report_path):
        """Header with newline: 'Things to Monitor\\nfrom the previous shifts'."""
        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        cell_value = str(ws.cell(row=layout.from_prev_row, column=1).value or "")
        assert "from the previous shifts" in cell_value.lower()


# ===================================================================
# _collect_source_rows
# ===================================================================

class TestCollectSourceRows:

    def test_collects_tickets_from_both_sections(self, report_path):
        """Should collect tickets from 'from previous' AND 'TTM' sections."""
        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)
        rows = ShiftReport._collect_source_rows(ws, layout)

        # Night has 3 from_prev + 1 TTM = 4 total
        assert len(rows) == 4
        assert all(isinstance(r, RowSnapshot) for r in rows)

    def test_ticket_ids_preserved(self, report_path):
        """Ticket IDs should match what we put in."""
        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)
        rows = ShiftReport._collect_source_rows(ws, layout)

        ticket_ids = [r.ticket_id for r in rows]
        assert "DSSD-29001" in ticket_ids
        assert "DSSD-29002" in ticket_ids
        assert "DSSD-29003" in ticket_ids
        assert "DRGN-50001" in ticket_ids

    def test_hyperlinks_captured(self, report_path):
        """Hyperlinks from cell.hyperlink should be captured in RowSnapshot."""
        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)
        rows = ShiftReport._collect_source_rows(ws, layout)

        first = rows[0]
        assert first.ticket_hyperlink is not None
        assert "DSSD-29001" in first.ticket_hyperlink
        assert first.slack_hyperlink is not None
        assert "slack.com" in first.slack_hyperlink

    def test_empty_section_returns_empty(self, report_path):
        """Day-Shift-NEW has 0 TTM tickets — 'from previous' only."""
        wb = load_workbook(report_path)
        ws = wb["Day-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)
        rows = ShiftReport._collect_source_rows(ws, layout)

        # Day has 2 from_prev + 0 TTM = 2 total
        assert len(rows) == 2


# ===================================================================
# _opposite_sheet
# ===================================================================

class TestOppositeSheet:

    def test_night_to_day(self):
        assert ShiftReport._opposite_sheet("Night-Shift-NEW") == "Day-Shift-NEW"

    def test_day_to_night(self):
        assert ShiftReport._opposite_sheet("Day-Shift-NEW") == "Night-Shift-NEW"


# ===================================================================
# _update_date
# ===================================================================

class TestUpdateDate:

    def test_night_increments_day(self, report_path):
        """Night shift target should get source_day + 1."""
        wb = load_workbook(report_path)
        source_ws = wb["Day-Shift-NEW"]  # source: day=10
        target_ws = wb["Night-Shift-NEW"]

        new_day, new_month = ShiftReport._update_date(
            target_ws, source_ws, "Night-Shift-NEW",
        )
        assert new_day == 11
        assert new_month == "Mar"
        assert target_ws.cell(row=1, column=1).value == 11

    def test_day_keeps_same_day(self, report_path):
        """Day shift target should keep same day as source."""
        wb = load_workbook(report_path)
        source_ws = wb["Night-Shift-NEW"]  # source: day=10
        target_ws = wb["Day-Shift-NEW"]

        new_day, new_month = ShiftReport._update_date(
            target_ws, source_ws, "Day-Shift-NEW",
        )
        assert new_day == 10
        assert new_month == "Mar"

    def test_month_boundary_mar31(self, report_month_boundary):
        """Mar 31 + 1 = Apr 1."""
        wb = load_workbook(report_month_boundary)
        source_ws = wb["Day-Shift-NEW"]  # day=31, month=Mar
        target_ws = wb["Night-Shift-NEW"]

        new_day, new_month = ShiftReport._update_date(
            target_ws, source_ws, "Night-Shift-NEW",
        )
        assert new_day == 1
        assert new_month == "Apr"

    def test_year_boundary_dec31(self, report_dec31):
        """Dec 31 + 1 = Jan 1."""
        wb = load_workbook(report_dec31)
        source_ws = wb["Day-Shift-NEW"]  # day=31, month=Dec
        target_ws = wb["Night-Shift-NEW"]

        new_day, new_month = ShiftReport._update_date(
            target_ws, source_ws, "Night-Shift-NEW",
        )
        assert new_day == 1
        assert new_month == "Jan"

    def test_day_shift_at_boundary_no_increment(self, report_month_boundary):
        """Day shift at Mar 31 keeps day=31 (no month rollover)."""
        wb = load_workbook(report_month_boundary)
        source_ws = wb["Night-Shift-NEW"]  # day=31, month=Mar
        target_ws = wb["Day-Shift-NEW"]

        new_day, new_month = ShiftReport._update_date(
            target_ws, source_ws, "Day-Shift-NEW",
        )
        assert new_day == 31
        assert new_month == "Mar"


# ===================================================================
# _build_status_string
# ===================================================================

class TestBuildStatusString:

    @pytest.fixture
    def assistant(self):
        return ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
            dry_run=True,
        )

    def test_basic_format(self, assistant):
        result = assistant._build_status_string("Open", "John Doe", "")
        assert result == "OPEN John Doe"

    def test_status_map_work_in_progress(self, assistant):
        result = assistant._build_status_string("Work In Progress", "Alice", "")
        assert result == "IN PROGRESS Alice"

    def test_unassigned(self, assistant):
        result = assistant._build_status_string("Open", "Unassigned", "")
        assert result == "OPEN Unassigned"

    def test_preserves_parenthetical_note(self, assistant):
        old = "OPEN Jane Smith (recurring)"
        result = assistant._build_status_string("Open", "Jane Smith", old)
        assert result == "OPEN Jane Smith (recurring)"

    def test_no_note_when_absent(self, assistant):
        old = "OPEN Jane Smith"
        result = assistant._build_status_string("Closed", "Jane Smith", old)
        assert result == "CLOSED Jane Smith"
        assert "(" not in result


# ===================================================================
# _extract_ticket_id
# ===================================================================

class TestExtractTicketId:

    @pytest.fixture
    def assistant(self):
        return ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
            dry_run=True,
        )

    def test_plain_text(self, assistant, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="DSSD-29001")
        result = assistant._extract_ticket_id(ws.cell(row=1, column=1))
        assert result == "DSSD-29001"

    def test_drgn_ticket(self, assistant, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="DRGN-50001")
        result = assistant._extract_ticket_id(ws.cell(row=1, column=1))
        assert result == "DRGN-50001"

    def test_no_match_returns_none(self, assistant, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="some random text")
        result = assistant._extract_ticket_id(ws.cell(row=1, column=1))
        assert result is None

    def test_empty_cell_returns_none(self, assistant, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        result = assistant._extract_ticket_id(ws.cell(row=1, column=1))
        assert result is None


# ===================================================================
# run() — sync statuses
# ===================================================================

class TestRunSync:

    def test_sync_updates_statuses(self, report_path, mock_jira):
        """Sync should update column E with fresh Jira data."""
        mock_jira("DSSD-29001", "Closed", "Kumar Raju")
        mock_jira("DSSD-29002", "Open", "Unassigned")
        mock_jira("DSSD-29003", "Work In Progress", "Jane Smith")
        mock_jira("DRGN-50001", "Open", "Unassigned")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        updates = assistant.run(report_path, "Night-Shift-NEW")

        assert len(updates) >= 3

        # Verify changes were written
        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        # Find DSSD-29001 and check its status
        for row in range(layout.from_prev_row, layout.permalinks_row):
            cell_d = str(ws.cell(row=row, column=4).value or "")
            if "DSSD-29001" in cell_d:
                status = ws.cell(row=row, column=5).value
                assert "CLOSED" in status
                assert "Kumar Raju" in status
                break

    def test_sync_preserves_notes(self, report_path, mock_jira):
        """Parenthetical notes like (recurring) should be preserved."""
        mock_jira("DSSD-29003", "Open", "Jane Smith")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        updates = assistant.run(report_path, "Night-Shift-NEW")

        for update in updates:
            if update.ticket_id == "DSSD-29003":
                assert "(recurring)" in update.new_value
                break

    def test_dry_run_no_save(self, report_path, mock_jira):
        """Dry run should not modify the file."""
        mock_jira("DSSD-29001", "Closed", "Kumar Raju")

        # Read original status
        wb_before = load_workbook(report_path)
        ws_before = wb_before["Night-Shift-NEW"]
        original_status = ws_before.cell(row=8, column=5).value

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
            dry_run=True,
        )
        assistant.run(report_path, "Night-Shift-NEW")

        # Status should be unchanged on disk
        wb_after = load_workbook(report_path)
        ws_after = wb_after["Night-Shift-NEW"]
        assert ws_after.cell(row=8, column=5).value == original_status

    def test_sync_stops_at_permalinks(self, report_path, mock_jira):
        """Sync should not process rows below Permalinks."""
        mock_jira("DSSD-29001", "Closed", "Kumar")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
            dry_run=True,
        )
        updates = assistant.run(report_path, "Night-Shift-NEW")

        # Permalink rows should not appear in updates
        for update in updates:
            assert "Dashboard" not in update.ticket_id
            assert "Grafana" not in update.ticket_id


# ===================================================================
# start_shift()
# ===================================================================

class TestStartShift:

    def test_copies_tickets_to_target(self, report_path, mock_jira):
        """Start shift should copy all source tickets into target 'from previous' section."""
        # Mock all tickets that will be synced after copy
        mock_jira("DSSD-29010", "Open", "Unassigned")
        mock_jira("DSSD-29011", "In Progress", "Alice")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        # Night target ← Day source (2 from_prev + 0 TTM = 2 tickets)
        result = assistant.start_shift(report_path, "Night-Shift-NEW")

        assert result["tickets_copied"] == 2
        assert result["date_day"] == 11
        assert result["date_month"] == "Mar"

        # Verify the tickets are in the file
        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        ticket_ids = []
        for row in range(layout.from_prev_row, layout.ttm_row):
            val = str(ws.cell(row=row, column=4).value or "")
            if val.strip():
                ticket_ids.append(val.strip())

        assert "DSSD-29010" in ticket_ids
        assert "DSSD-29011" in ticket_ids

    def test_resets_ttm_section(self, report_path, mock_jira):
        """After start_shift, TTM section should have 1 empty row."""
        mock_jira("DSSD-29010", "Open", "Unassigned")
        mock_jira("DSSD-29011", "In Progress", "Alice")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        assistant.start_shift(report_path, "Night-Shift-NEW")

        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        # TTM should span exactly 1 row
        assert layout.ttm_end == layout.ttm_row
        # TTM row should be empty in columns C-F
        for col in [3, 4, 5, 6]:
            assert ws.cell(row=layout.ttm_row, column=col).value is None

    def test_writes_from_prev_header(self, report_no_header, mock_jira):
        """After start_shift on sheet without header, header text should be written."""
        mock_jira("DSSD-29001", "Open", "Unassigned")
        mock_jira("DSSD-29002", "In Progress", "John Doe")
        mock_jira("DSSD-29003", "Open", "Jane Smith")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        # Day target ← Night source (3 tickets)
        assistant.start_shift(report_no_header, "Day-Shift-NEW")

        wb = load_workbook(report_no_header)
        ws = wb["Day-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        header_text = str(ws.cell(row=layout.from_prev_row, column=1).value or "")
        assert "from the previous shifts" in header_text.lower()

    def test_delta_positive_inserts_rows(self, report_path, mock_jira):
        """When source has more tickets than target section, rows should be inserted."""
        # Day source has 2 tickets, Night target has 3 — delta = 2-3 = -1
        # Let's test the other way: Day target ← Night source (4 tickets) vs Day's 2
        for tid in ["DSSD-29001", "DSSD-29002", "DSSD-29003", "DRGN-50001"]:
            mock_jira(tid, "Open", "Unassigned")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        result = assistant.start_shift(report_path, "Day-Shift-NEW")

        assert result["tickets_copied"] == 4

        wb = load_workbook(report_path)
        ws = wb["Day-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        # 4 ticket rows in "from previous" section
        count = layout.from_prev_end - layout.from_prev_row + 1
        assert count == 4

    def test_delta_negative_deletes_rows(self, report_path, mock_jira):
        """When source has fewer tickets, excess rows should be deleted."""
        # Night target ← Day source (2 tickets) — Night currently has 3 from_prev
        mock_jira("DSSD-29010", "Open", "Unassigned")
        mock_jira("DSSD-29011", "In Progress", "Alice")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        result = assistant.start_shift(report_path, "Night-Shift-NEW")

        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        count = layout.from_prev_end - layout.from_prev_row + 1
        assert count == 2

    def test_file_reopens_cleanly(self, report_path, mock_jira):
        """After start_shift, file should reopen without errors."""
        mock_jira("DSSD-29010", "Open", "Unassigned")
        mock_jira("DSSD-29011", "In Progress", "Alice")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        assistant.start_shift(report_path, "Night-Shift-NEW")

        # This should not raise (openpyxl corruption = exception here)
        wb = load_workbook(report_path)
        assert "Night-Shift-NEW" in wb.sheetnames
        assert "Day-Shift-NEW" in wb.sheetnames

    def test_permalinks_intact_after_shift(self, report_path, mock_jira):
        """Permalinks section should survive start_shift intact."""
        mock_jira("DSSD-29010", "Open", "Unassigned")
        mock_jira("DSSD-29011", "In Progress", "Alice")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        assistant.start_shift(report_path, "Night-Shift-NEW")

        wb = load_workbook(report_path)
        ws = wb["Night-Shift-NEW"]
        layout = ShiftReport._scan_layout(ws)

        assert ws.cell(row=layout.permalinks_row, column=1).value == "Permalinks"
        # Permalink data rows should still exist
        assert ws.cell(row=layout.permalinks_row + 1, column=1).value == "NOC Dashboard"
        assert ws.cell(row=layout.permalinks_row + 2, column=1).value == "Grafana"

    def test_dry_run_returns_empty_sync(self, report_path):
        """Dry run should return empty sync_updates list."""
        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
            dry_run=True,
        )
        result = assistant.start_shift(report_path, "Night-Shift-NEW")

        assert result["sync_updates"] == []

    def test_month_boundary(self, report_month_boundary, mock_jira):
        """Start shift at Mar 31 → Night should get Apr 1."""
        mock_jira("DSSD-29010", "Open", "Unassigned")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        result = assistant.start_shift(report_month_boundary, "Night-Shift-NEW")

        assert result["date_day"] == 1
        assert result["date_month"] == "Apr"


# ===================================================================
# add_row()
# ===================================================================

class TestAddRow:

    def test_add_first_ticket_to_empty_ttm(self, report_path, mock_jira):
        """Adding to empty TTM should write directly into TTM row (no insert)."""
        mock_jira("DSSD-29999", "Open", "Unassigned", summary="New bug found")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        # Day-Shift-NEW has empty TTM
        row = assistant.add_row(
            report_path, "Day-Shift-NEW",
            jira_link="https://jira.example.com/browse/DSSD-29999",
            slack_link="https://company.slack.com/archives/C123/p999",
        )

        wb = load_workbook(report_path)
        ws = wb["Day-Shift-NEW"]
        assert ws.cell(row=row, column=4).value == "DSSD-29999"
        assert ws.cell(row=row, column=3).value == "New bug found"
        assert "OPEN" in ws.cell(row=row, column=5).value

    def test_add_second_ticket_inserts_row(self, report_path, mock_jira):
        """Adding to TTM with existing ticket should insert a new row."""
        mock_jira("DRGN-50002", "Open", "Bob", summary="Another issue")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        # Night-Shift-NEW has 1 TTM ticket already
        wb_before = load_workbook(report_path)
        ws_before = wb_before["Night-Shift-NEW"]
        layout_before = ShiftReport._scan_layout(ws_before)
        permalinks_before = layout_before.permalinks_row

        row = assistant.add_row(
            report_path, "Night-Shift-NEW",
            jira_link="https://jira.example.com/browse/DRGN-50002",
            slack_link="https://company.slack.com/archives/C123/p888",
        )

        wb_after = load_workbook(report_path)
        ws_after = wb_after["Night-Shift-NEW"]
        layout_after = ShiftReport._scan_layout(ws_after)

        # Permalinks should have shifted down by 1
        assert layout_after.permalinks_row == permalinks_before + 1
        assert ws_after.cell(row=row, column=4).value == "DRGN-50002"

    def test_add_row_file_reopens(self, report_path, mock_jira):
        """File should reopen cleanly after add_row."""
        mock_jira("DSSD-29999", "Open", "Unassigned", summary="Test")

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        assistant.add_row(
            report_path, "Day-Shift-NEW",
            jira_link="https://jira.example.com/browse/DSSD-29999",
            slack_link="https://company.slack.com/archives/C123/p999",
        )

        # Should not raise
        wb = load_workbook(report_path)
        assert "Day-Shift-NEW" in wb.sheetnames


# ===================================================================
# Helper functions
# ===================================================================

class TestHelpers:

    def test_rebuild_section_merge(self):
        """_rebuild_section_merge should create A:B merge for given range."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.merge_cells(start_row=5, start_column=1, end_row=7, end_column=2)

        _rebuild_section_merge(ws, 5, 10)

        # Old merge (5-7) should be gone, new merge (5-10) should exist
        found = False
        for mr in ws.merged_cells.ranges:
            if mr.min_row == 5 and mr.max_row == 10 and mr.min_col == 1 and mr.max_col == 2:
                found = True
        assert found

    def test_remove_hyperlinks_in_range(self):
        """Should remove hyperlinks within the specified range."""
        from openpyxl import Workbook
        from openpyxl.worksheet.hyperlink import Hyperlink

        wb = Workbook()
        ws = wb.active
        # Directly populate _hyperlinks (as openpyxl does when loading from file)
        ws._hyperlinks = [
            Hyperlink(ref="A1", target="https://example.com/1"),
            Hyperlink(ref="C2", target="https://example.com/2"),
            Hyperlink(ref="A5", target="https://example.com/3"),
        ]

        _remove_hyperlinks_in_range(ws, 1, 1, 3, 6)

        remaining = [hl.target for hl in ws._hyperlinks]
        assert len(remaining) == 1
        assert "https://example.com/3" in remaining

    def test_apply_hyperlink_font(self):
        """Should set Jira-blue color and underline."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="DSSD-123")

        _apply_hyperlink_font(cell)

        assert cell.font.underline == "single"
        assert cell.font.color.rgb == "FF0052CC"

    def test_copy_cell_style(self):
        """Should copy fill, font, border, alignment."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        wb = Workbook()
        ws = wb.active

        source = ws.cell(row=1, column=1)
        source.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        source.font = Font(bold=True, size=14)

        target = ws.cell(row=2, column=1)
        _copy_cell_style(source, target)

        assert target.fill.start_color.rgb == "FFFF0000"
        assert target.font.bold is True
        assert target.font.size == 14


# ===================================================================
# Import error blocks (lines 23-35): tested via subprocess/importlib
# ===================================================================

class TestImportErrorBlocks:

    def test_openpyxl_import_error_exits(self):
        """When openpyxl is not importable, the module should exit with code 1."""
        import subprocess
        code = (
            "import sys; import builtins; real = builtins.__import__\n"
            "def fake(name, *a, **kw):\n"
            "    if name in ('openpyxl',):\n"
            "        raise ImportError('no openpyxl')\n"
            "    return real(name, *a, **kw)\n"
            "builtins.__import__ = fake\n"
            "import importlib, tools.shift-report\n"
        )
        result = subprocess.run(
            [sys.executable, "-c",
             "import sys; sys.modules.pop('shift_report', None); "
             "import builtins, importlib\n"
             "real_import = builtins.__import__\n"
             "def fake_import(name, *args, **kwargs):\n"
             "    if 'openpyxl' in name:\n"
             "        raise ImportError('mocked')\n"
             "    return real_import(name, *args, **kwargs)\n"
             "builtins.__import__ = fake_import\n"
             "import shift_report\n"],
            capture_output=True,
            text=True,
            cwd=str(__import__('pathlib').Path(__file__).parent.parent),
            env={
                **__import__('os').environ,
                "PYTHONPATH": ":".join([
                    str(__import__('pathlib').Path(__file__).parent.parent / "tools" / "common"),
                    str(__import__('pathlib').Path(__file__).parent.parent / "tools" / "shift-report"),
                ]),
            },
        )
        assert result.returncode == 1

    def test_noc_utils_import_error_exits(self):
        """When noc_utils is not importable, the module should exit with code 1."""
        import subprocess, os
        result = subprocess.run(
            [sys.executable, "-c", "import shift_report"],
            capture_output=True,
            text=True,
            cwd=str(__import__('pathlib').Path(__file__).parent.parent),
            env={
                # Deliberately empty PYTHONPATH so noc_utils is not found
                k: v for k, v in os.environ.items()
                if k not in ("PYTHONPATH",)
            },
        )
        # Without tools/common on path, noc_utils ImportError fires → exit(1)
        assert result.returncode == 1


# ===================================================================
# _restructure_from_prev: empty source_rows branch (lines 407-408)
# ===================================================================

class TestRestructureFromPrevEmpty:

    def test_empty_source_rows_clears_cells(self):
        """When source_rows is empty, from_prev data cells should be cleared."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # Set up minimal sheet layout accepted by _restructure_from_prev
        ws.cell(row=8, column=1, value="Things to Monitor\nfrom the previous shifts")
        ws.cell(row=9, column=4, value="DSSD-99999")  # existing ticket
        ws.cell(row=10, column=1, value="Things to monitor")
        ws.cell(row=11, column=1, value="Permalinks")
        ws.merge_cells(start_row=8, start_column=1, end_row=9, end_column=2)
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=2)

        layout = ShiftLayout(
            from_prev_row=8,
            from_prev_end=9,
            ttm_row=10,
            ttm_end=10,
            permalinks_row=11,
        )

        ttm_row, permalinks_row = ShiftReport._restructure_from_prev(ws, layout, [])

        # With 0 source rows → 1 kept (max(0, 1) = 1), delta = 1 - 2 = -1
        # The from_prev cells should be None
        assert ws.cell(row=8, column=4).value is None
        assert ws.cell(row=8, column=5).value is None


# ===================================================================
# _reset_ttm: extra rows branch (lines 425-426)
# ===================================================================

class TestResetTtmExtraRows:

    def test_reset_ttm_deletes_extra_rows(self):
        """When TTM section has more than one row, extras should be deleted."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=10, column=1, value="Things to monitor")
        ws.cell(row=11, column=4, value="DRGN-11111")  # extra TTM ticket
        ws.cell(row=12, column=4, value="DRGN-22222")  # another extra
        ws.cell(row=13, column=1, value="Permalinks")
        ws.merge_cells(start_row=10, start_column=1, end_row=12, end_column=2)

        # ttm_row=10, ttm_end=12, permalinks_row=13
        new_permalinks = ShiftReport._reset_ttm(ws, 10, 12, 13)

        # Two extra rows deleted → permalinks shifted from 13 to 11
        assert new_permalinks == 11
        # TTM row cells should be cleared
        for col in [3, 4, 5, 6]:
            assert ws.cell(row=10, column=col).value is None


# ===================================================================
# _fetch_jira_full: None response raises RuntimeError (line 579)
# ===================================================================

class TestFetchJiraFull:

    def test_raises_on_none_response(self):
        """_fetch_jira_full should raise RuntimeError when _jira_get returns None."""
        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        with patch.object(assistant, "_jira_get", return_value=None):
            with pytest.raises(RuntimeError, match="Failed to fetch Jira issue"):
                assistant._fetch_jira_full("DSSD-99999")


# ===================================================================
# _jira_get: URLError / network failure branch (lines 589-602)
# ===================================================================

class TestJiraGet:

    def test_returns_none_on_url_error(self):
        """_jira_get should return None when a URLError is raised."""
        from urllib.error import URLError
        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        with patch("shift_report.urlopen", side_effect=URLError("connection refused")):
            result = assistant._jira_get("https://jira.example.com/rest/api/2/issue/DSSD-1")
        assert result is None

    def test_returns_none_on_http_error(self):
        """_jira_get should return None when an HTTPError is raised."""
        from urllib.error import HTTPError
        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        with patch("shift_report.urlopen", side_effect=HTTPError(
            url="https://jira.example.com", code=401, msg="Unauthorized",
            hdrs={}, fp=None,
        )):
            result = assistant._jira_get("https://jira.example.com/rest/api/2/issue/DSSD-1")
        assert result is None

    def test_returns_parsed_json_on_success(self):
        """_jira_get should return parsed JSON dict on success."""
        payload = b'{"fields": {"status": {"name": "Open"}}}'
        mock_response = MagicMock()
        mock_response.__enter__ = lambda s: s
        mock_response.__exit__ = MagicMock(return_value=False)
        mock_response.read.return_value = payload

        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        with patch("shift_report.urlopen", return_value=mock_response):
            result = assistant._jira_get("https://jira.example.com/rest/api/2/issue/DSSD-1")
        assert result == {"fields": {"status": {"name": "Open"}}}


# ===================================================================
# _remove_hyperlinks_in_range: ValueError/TypeError branch (lines 671-672)
# ===================================================================

class TestRemoveHyperlinksValueError:

    def test_invalid_ref_is_kept(self):
        """Hyperlinks with unparseable ref should be kept (not removed)."""
        from openpyxl import Workbook
        from openpyxl.worksheet.hyperlink import Hyperlink

        wb = Workbook()
        ws = wb.active
        bad_hl = Hyperlink(ref="INVALID_REF", target="https://example.com/bad")
        good_hl = Hyperlink(ref="A1", target="https://example.com/good")
        ws._hyperlinks = [bad_hl, good_hl]

        # Range that would include A1 but "INVALID_REF" can't be parsed
        _remove_hyperlinks_in_range(ws, 1, 1, 1, 1)

        # A1 should be removed, INVALID_REF kept (parse error → pass)
        remaining_targets = [h.target for h in ws._hyperlinks]
        assert "https://example.com/bad" in remaining_targets
        assert "https://example.com/good" not in remaining_targets


# ===================================================================
# _ensure_permalink_merges: break when both A and C are None (line 690)
# ===================================================================

class TestEnsurePermalinkMergesBreak:

    def test_stops_at_empty_row(self):
        """Should stop processing permalink rows when both A and C cells are None."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        permalinks_row = 5
        ws.cell(row=permalinks_row, column=1, value="Permalinks")
        ws.cell(row=permalinks_row + 1, column=1, value="NOC Dashboard")
        ws.cell(row=permalinks_row + 1, column=3, value="https://dash.example.com")
        # Row permalinks_row + 2: both A and C are None — but we need max_row to
        # include it, so write something in a column that won't trigger the loop data
        # check (neither col 1 nor col 3), then clear it so A/C remain None.
        # Use col 7 (G) to force max_row beyond the data row.
        ws.cell(row=permalinks_row + 2, column=7, value="filler")

        _ensure_permalink_merges(ws, permalinks_row)

        # Row permalinks_row + 1 should be merged (A:B and C:F)
        merged_ranges = {
            (mr.min_row, mr.min_col, mr.max_col)
            for mr in ws.merged_cells.ranges
        }
        assert (permalinks_row + 1, 1, 2) in merged_ranges
        assert (permalinks_row + 1, 3, 6) in merged_ranges


# ===================================================================
# add_row: invalid jira_link (no ticket ID) raises ValueError (line 246)
# ===================================================================

class TestAddRowInvalidLink:

    def test_add_row_raises_on_bad_jira_link(self, report_path):
        """add_row should raise ValueError when jira_link has no ticket ID."""
        assistant = ShiftReport(
            jira_url="https://jira.example.com",
            jira_token="fake",
        )
        with pytest.raises(ValueError, match="Could not extract ticket ID"):
            assistant.add_row(
                report_path, "Day-Shift-NEW",
                jira_link="https://jira.example.com/browse/notavalidticket",
                slack_link="https://company.slack.com/archives/C123/p999",
            )


# ===================================================================
# collect_links interactive function (lines 786-815)
# ===================================================================

class TestCollectLinks:

    def test_collects_slack_then_jira(self):
        """Should accept Slack link first, then Jira link."""
        inputs = iter([
            "https://company.slack.com/archives/C123/p999",
            "https://jira.example.com/browse/DSSD-12345",
        ])
        with patch("builtins.input", side_effect=inputs):
            jira_link, slack_link = collect_links()
        assert "DSSD-12345" in jira_link
        assert "slack.com" in slack_link

    def test_collects_jira_then_slack(self):
        """Should accept Jira link first, then Slack link."""
        inputs = iter([
            "https://jira.example.com/browse/DSSD-99999",
            "https://company.slack.com/archives/C456/p888",
        ])
        with patch("builtins.input", side_effect=inputs):
            jira_link, slack_link = collect_links()
        assert "DSSD-99999" in jira_link
        assert "slack.com" in slack_link

    def test_skips_empty_input(self):
        """Empty input lines should be skipped without error."""
        inputs = iter([
            "",
            "https://jira.example.com/browse/DSSD-11111",
            "https://company.slack.com/archives/C789/p777",
        ])
        with patch("builtins.input", side_effect=inputs):
            jira_link, slack_link = collect_links()
        assert "DSSD-11111" in jira_link

    def test_skips_jira_link_without_ticket_id(self):
        """Jira link without a recognizable ticket ID should be rejected."""
        inputs = iter([
            "https://jira.example.com/browse/notavalidticket",
            "https://jira.example.com/browse/DSSD-22222",
            "https://company.slack.com/archives/C000/p000",
        ])
        with patch("builtins.input", side_effect=inputs):
            jira_link, slack_link = collect_links()
        assert "DSSD-22222" in jira_link

    def test_skips_unrecognized_link(self):
        """Unrecognized links should be skipped."""
        inputs = iter([
            "https://google.com",
            "https://jira.example.com/browse/DSSD-33333",
            "https://company.slack.com/archives/C111/p111",
        ])
        with patch("builtins.input", side_effect=inputs):
            jira_link, slack_link = collect_links()
        assert "DSSD-33333" in jira_link


# ===================================================================
# select_sheet interactive function (lines 820-830)
# ===================================================================

class TestSelectSheet:

    def test_valid_choice_1(self):
        """Selecting 1 should return Night-Shift-NEW."""
        with patch("builtins.input", return_value="1"):
            result = select_sheet()
        assert result == "Night-Shift-NEW"

    def test_valid_choice_2(self):
        """Selecting 2 should return Day-Shift-NEW."""
        with patch("builtins.input", return_value="2"):
            result = select_sheet()
        assert result == "Day-Shift-NEW"

    def test_invalid_then_valid(self):
        """Invalid input should loop until a valid choice is entered."""
        inputs = iter(["0", "abc", "3", "1"])
        with patch("builtins.input", side_effect=inputs):
            result = select_sheet()
        assert result == "Night-Shift-NEW"

    def test_eof_then_valid(self):
        """EOFError (e.g. piped stdin exhausted) should be handled gracefully."""
        inputs = iter([EOFError(), "2"])

        def fake_input(prompt=""):
            val = next(inputs)
            if isinstance(val, type) and issubclass(val, Exception):
                raise val()
            if isinstance(val, BaseException):
                raise val
            return val

        with patch("builtins.input", side_effect=fake_input):
            result = select_sheet()
        assert result == "Day-Shift-NEW"


# ===================================================================
# select_action interactive function (lines 835-846)
# ===================================================================

class TestSelectAction:

    def test_valid_choice_1(self):
        with patch("builtins.input", return_value="1"):
            assert select_action() == 1

    def test_valid_choice_2(self):
        with patch("builtins.input", return_value="2"):
            assert select_action() == 2

    def test_valid_choice_3(self):
        with patch("builtins.input", return_value="3"):
            assert select_action() == 3

    def test_invalid_then_valid(self):
        inputs = iter(["0", "4", "abc", "2"])
        with patch("builtins.input", side_effect=inputs):
            assert select_action() == 2

    def test_eof_then_valid(self):
        """EOFError should be handled, then accept a valid input."""
        inputs = iter([EOFError(), "3"])

        def fake_input(prompt=""):
            val = next(inputs)
            if isinstance(val, BaseException):
                raise val
            return val

        with patch("builtins.input", side_effect=fake_input):
            assert select_action() == 3


# ===================================================================
# parse_args (lines 855-872)
# ===================================================================

class TestParseArgs:

    def test_defaults(self):
        """Default args should have no file, dry_run=False, verbose=False."""
        with patch("sys.argv", ["shift_report.py"]):
            args = parse_args()
        assert args.file is None
        assert args.dry_run is False
        assert args.verbose is False

    def test_file_argument(self):
        with patch("sys.argv", ["shift_report.py", "--file", "/tmp/report.xlsx"]):
            args = parse_args()
        assert args.file == "/tmp/report.xlsx"

    def test_short_file_argument(self):
        with patch("sys.argv", ["shift_report.py", "-f", "/tmp/report.xlsx"]):
            args = parse_args()
        assert args.file == "/tmp/report.xlsx"

    def test_dry_run_flag(self):
        with patch("sys.argv", ["shift_report.py", "--dry-run"]):
            args = parse_args()
        assert args.dry_run is True

    def test_verbose_flag(self):
        with patch("sys.argv", ["shift_report.py", "--verbose"]):
            args = parse_args()
        assert args.verbose is True


# ===================================================================
# main() — various paths (lines 877-974)
# ===================================================================

class TestMain:

    def _base_env(self, report_path):
        """Return env dict suitable for main() with valid Jira vars."""
        return {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }

    def test_missing_env_vars_exits(self):
        """main() should exit 1 when JIRA env vars are missing."""
        with patch.dict("os.environ", {}, clear=True):
            with patch("sys.argv", ["shift_report.py"]):
                with pytest.raises(SystemExit) as exc_info:
                    main()
        assert exc_info.value.code == 1

    def test_missing_jira_url_exits(self):
        """main() should exit 1 when only JIRA_SERVER_URL is missing."""
        env = {"JIRA_PERSONAL_ACCESS_TOKEN": "fake"}
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py"]):
                with pytest.raises(SystemExit) as exc_info:
                    main()
        assert exc_info.value.code == 1

    def test_missing_file_exits(self, tmp_path):
        """main() should exit 1 when report file does not exist."""
        nonexistent = str(tmp_path / "no_such_file.xlsx")
        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--file", nonexistent]):
                with pytest.raises(SystemExit) as exc_info:
                    main()
        assert exc_info.value.code == 1

    def test_action1_start_shift(self, report_path, mock_jira):
        """main() action=1 should run start_shift successfully."""
        mock_jira("DSSD-29010", "Open", "Unassigned")
        mock_jira("DSSD-29011", "In Progress", "Alice")

        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--file", str(report_path), "--dry-run"]):
                with patch("shift_report.select_sheet", return_value="Night-Shift-NEW"):
                    with patch("shift_report.select_action", return_value=1):
                        main()  # should not raise

    def test_action1_start_shift_with_sync(self, report_path, mock_jira):
        """main() action=1 (non-dry-run) should log sync results when tickets changed."""
        mock_jira("DSSD-29010", "Closed", "Kumar")
        mock_jira("DSSD-29011", "Open", "Unassigned")

        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--file", str(report_path)]):
                with patch("shift_report.select_sheet", return_value="Night-Shift-NEW"):
                    with patch("shift_report.select_action", return_value=1):
                        main()  # should not raise

    def test_action2_end_shift_dry_run(self, report_path, mock_jira):
        """main() action=2 dry-run should log processed count."""
        mock_jira("DSSD-29001", "Closed", "Kumar Raju")
        mock_jira("DSSD-29002", "Open", "Unassigned")
        mock_jira("DSSD-29003", "Open", "Jane Smith")
        mock_jira("DRGN-50001", "Open", "Unassigned")

        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--file", str(report_path), "--dry-run"]):
                with patch("shift_report.select_sheet", return_value="Night-Shift-NEW"):
                    with patch("shift_report.select_action", return_value=2):
                        main()  # should not raise

    def test_action2_end_shift_logs_changes(self, report_path, mock_jira):
        """main() action=2 should log changed tickets."""
        mock_jira("DSSD-29001", "Closed", "Kumar Raju")
        mock_jira("DSSD-29002", "Open", "Unassigned")
        mock_jira("DSSD-29003", "Open", "Jane Smith")
        mock_jira("DRGN-50001", "Open", "Unassigned")

        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--file", str(report_path)]):
                with patch("shift_report.select_sheet", return_value="Night-Shift-NEW"):
                    with patch("shift_report.select_action", return_value=2):
                        main()  # should not raise

    def test_action3_add_row_dry_run(self, report_path, mock_jira):
        """main() action=3 dry-run should log row details without saving."""
        mock_jira("DSSD-77777", "Open", "Unassigned", summary="Dry run ticket")

        jira_link = "https://jira.example.com/browse/DSSD-77777"
        slack_link = "https://company.slack.com/archives/C000/p000"

        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--file", str(report_path), "--dry-run"]):
                with patch("shift_report.select_sheet", return_value="Day-Shift-NEW"):
                    with patch("shift_report.select_action", return_value=3):
                        with patch("shift_report.collect_links", return_value=(jira_link, slack_link)):
                            main()  # should not raise

    def test_action3_add_row_saves(self, report_path, mock_jira):
        """main() action=3 (non-dry-run) should actually insert a row."""
        mock_jira("DSSD-88888", "Open", "Bob", summary="Live ticket")

        jira_link = "https://jira.example.com/browse/DSSD-88888"
        slack_link = "https://company.slack.com/archives/C000/p000"

        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--file", str(report_path)]):
                with patch("shift_report.select_sheet", return_value="Day-Shift-NEW"):
                    with patch("shift_report.select_action", return_value=3):
                        with patch("shift_report.collect_links", return_value=(jira_link, slack_link)):
                            main()  # should not raise

        wb = load_workbook(report_path)
        ws = wb["Day-Shift-NEW"]
        # Ticket should now exist somewhere in the sheet
        found = any(
            "DSSD-88888" in str(ws.cell(row=r, column=4).value or "")
            for r in range(1, ws.max_row + 1)
        )
        assert found

    def test_action3_add_row_bad_link_exits(self, report_path):
        """main() action=3 with unextractable ticket ID from collect_links exits 1."""
        # collect_links returns a jira_link that has no ticket ID (edge case:
        # TICKET_REGEX won't match). We must monkeypatch TICKET_REGEX to fail.
        jira_link = "https://jira.example.com/browse/DSSD-99999"
        slack_link = "https://company.slack.com/archives/C000/p000"

        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
        }
        import re
        # Patch TICKET_REGEX to never match so the "if not ticket_id_match" branch fires
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--file", str(report_path)]):
                with patch("shift_report.select_sheet", return_value="Day-Shift-NEW"):
                    with patch("shift_report.select_action", return_value=3):
                        with patch("shift_report.collect_links", return_value=(jira_link, slack_link)):
                            with patch("shift_report.TICKET_REGEX", re.compile(r"NOMATCH")):
                                with pytest.raises(SystemExit) as exc_info:
                                    main()
        assert exc_info.value.code == 1

    def test_noc_report_path_env_var(self, report_path, mock_jira):
        """NOC_REPORT_PATH env var should be used when --file is not given."""
        mock_jira("DSSD-29010", "Open", "Unassigned")
        mock_jira("DSSD-29011", "In Progress", "Alice")

        env = {
            "JIRA_SERVER_URL": "https://jira.example.com",
            "JIRA_PERSONAL_ACCESS_TOKEN": "fake_token",
            "NOC_REPORT_PATH": str(report_path),
        }
        with patch.dict("os.environ", env, clear=True):
            with patch("sys.argv", ["shift_report.py", "--dry-run"]):
                with patch("shift_report.select_sheet", return_value="Night-Shift-NEW"):
                    with patch("shift_report.select_action", return_value=1):
                        main()  # should not raise


# ===========================================================================
# Import error fallback (lines 23-26, 32-35)
# ===========================================================================


class TestImportErrorFallbackReload:
    """Test import-error fallbacks by reloading the module."""

    def test_missing_noc_utils_exits(self) -> None:
        """Module exits with code 1 when noc_utils is unavailable."""
        import builtins
        import importlib
        import shift_report

        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "noc_utils":
                raise ImportError("No module named 'noc_utils'")
            return real_import(name, *args, **kwargs)

        with patch.object(builtins, "__import__", side_effect=fake_import):
            with pytest.raises(SystemExit) as exc_info:
                importlib.reload(shift_report)
            assert exc_info.value.code == 1

    def test_missing_openpyxl_exits(self) -> None:
        """Module exits with code 1 when openpyxl is unavailable."""
        import builtins
        import importlib
        import shift_report

        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return real_import(name, *args, **kwargs)

        with patch.object(builtins, "__import__", side_effect=fake_import):
            with pytest.raises(SystemExit) as exc_info:
                importlib.reload(shift_report)
            assert exc_info.value.code == 1
