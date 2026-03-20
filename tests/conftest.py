"""Shared fixtures for NOC Toolkit tests."""

import sys
from pathlib import Path

import pytest

# Make tool modules importable
TOOLS_DIR = Path(__file__).resolve().parent.parent / "tools"
sys.path.insert(0, str(TOOLS_DIR / "shift-report"))
sys.path.insert(0, str(TOOLS_DIR / "pd-escalate"))
sys.path.insert(0, str(TOOLS_DIR / "freshness"))
sys.path.insert(0, str(TOOLS_DIR / "pd-merge"))
sys.path.insert(0, str(TOOLS_DIR / "pd-jobs"))
sys.path.insert(0, str(TOOLS_DIR / "pd-sync"))
sys.path.insert(0, str(TOOLS_DIR / "pd-monitor"))
sys.path.insert(0, str(TOOLS_DIR / "pd-resolve"))
sys.path.insert(0, str(TOOLS_DIR / "ticket-watch"))


# ---------------------------------------------------------------------------
# Excel fixture: builds a minimal two-sheet report matching real structure
# ---------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Reusable style constants
_ORANGE_FILL = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
_GRAY_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
_WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
_BOLD_FONT = Font(bold=True)


def _build_sheet(
    ws,
    sheet_label: str,
    day: int,
    month: str,
    time_range: str,
    from_prev_tickets: list,
    ttm_tickets: list,
    include_from_prev_header: bool = True,
):
    """Populate a worksheet with the standard NOC report structure.

    Each ticket is a dict: {summary, ticket_id, jira_url, status, slack_text, slack_url}
    """
    # Row 1: day | "Shift report" | "LiveNation"
    ws.cell(row=1, column=1, value=day)
    ws.cell(row=1, column=2, value="Shift report")
    ws.cell(row=1, column=3, value="LiveNation")

    # Row 2: month | time range
    ws.cell(row=2, column=1, value=month)
    ws.cell(row=2, column=2, value=time_range)

    # Row 3: headers
    for col, header in enumerate(["Topic", "", "Cause/Ticket", "Ticket", "Status", "Slack"], 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = _BOLD_FONT

    # Rows 4-7: static rows
    static_rows = ["Feeds delays", "Downtime", "Delays in Legacy", "Delays in DBKS"]
    for idx, label in enumerate(static_rows):
        row = 4 + idx
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=3, value="OK")

    current_row = 8  # TICKET_START_ROW

    # "Things to Monitor from the previous shifts" section
    if include_from_prev_header:
        ws.cell(row=current_row, column=1, value="Things to Monitor\nfrom the previous shifts")
    from_prev_start = current_row

    for ticket in from_prev_tickets:
        _write_test_ticket(ws, current_row, ticket)
        current_row += 1

    if not from_prev_tickets and include_from_prev_header:
        current_row += 1  # keep at least the header row

    from_prev_end = current_row - 1

    # Merge A:B for "from previous shifts" section
    if from_prev_end >= from_prev_start:
        ws.merge_cells(
            start_row=from_prev_start, start_column=1,
            end_row=from_prev_end, end_column=2,
        )
    ws.cell(row=from_prev_start, column=1).fill = _ORANGE_FILL

    # "Things to monitor" section
    ttm_start = current_row
    ws.cell(row=current_row, column=1, value="Things to monitor        ")

    for ticket in ttm_tickets:
        _write_test_ticket(ws, current_row, ticket)
        current_row += 1

    if not ttm_tickets:
        current_row += 1  # keep at least the header row

    ttm_end = current_row - 1

    # Merge A:B for "Things to monitor" section
    if ttm_end >= ttm_start:
        ws.merge_cells(
            start_row=ttm_start, start_column=1,
            end_row=ttm_end, end_column=2,
        )
    ws.cell(row=ttm_start, column=1).fill = _ORANGE_FILL

    # "Permalinks" row
    permalinks_row = current_row
    ws.cell(row=permalinks_row, column=1, value="Permalinks")
    ws.cell(row=permalinks_row, column=1).fill = _ORANGE_FILL
    ws.merge_cells(
        start_row=permalinks_row, start_column=1,
        end_row=permalinks_row, end_column=6,
    )

    # Two permalink rows
    for offset, (label, url) in enumerate([
        ("NOC Dashboard", "https://jira.example.com/dashboard"),
        ("Grafana", "https://grafana.example.com/status"),
    ], 1):
        row = permalinks_row + offset
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=3, value=url)
        ws.cell(row=row, column=1).fill = _GRAY_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)

    # White fill on cols G-H for all data rows (mimic real file)
    for row in range(1, permalinks_row + 3):
        for col in [7, 8]:
            ws.cell(row=row, column=col).fill = _WHITE_FILL


def _write_test_ticket(ws, row: int, ticket: dict) -> None:
    """Write a single ticket into a row."""
    ws.cell(row=row, column=3, value=ticket.get("summary", ""))
    ws.cell(row=row, column=3).fill = _GRAY_FILL

    cell_d = ws.cell(row=row, column=4, value=ticket.get("ticket_id", ""))
    cell_d.fill = _GRAY_FILL
    jira_url = ticket.get("jira_url")
    if jira_url:
        cell_d.hyperlink = jira_url

    ws.cell(row=row, column=5, value=ticket.get("status", ""))
    ws.cell(row=row, column=5).fill = _GRAY_FILL

    cell_f = ws.cell(row=row, column=6, value=ticket.get("slack_text", "slack_link"))
    cell_f.fill = _GRAY_FILL
    slack_url = ticket.get("slack_url")
    if slack_url:
        cell_f.hyperlink = slack_url


# ---------------------------------------------------------------------------
# Sample ticket data
# ---------------------------------------------------------------------------

SAMPLE_TICKETS_A = [
    {
        "summary": "Data export job failed",
        "ticket_id": "DSSD-29001",
        "jira_url": "https://jira.example.com/browse/DSSD-29001",
        "status": "OPEN Unassigned",
        "slack_text": "slack_link",
        "slack_url": "https://company.slack.com/archives/C123/p111",
    },
    {
        "summary": "Databricks batch job failing",
        "ticket_id": "DSSD-29002",
        "jira_url": "https://jira.example.com/browse/DSSD-29002",
        "status": "IN PROGRESS John Doe",
        "slack_text": "slack_link",
        "slack_url": "https://company.slack.com/archives/C123/p222",
    },
    {
        "summary": "RDS export failed to start",
        "ticket_id": "DSSD-29003",
        "jira_url": "https://jira.example.com/browse/DSSD-29003",
        "status": "OPEN Jane Smith (recurring)",
        "slack_text": "slack_link",
        "slack_url": "https://company.slack.com/archives/C123/p333",
    },
]

SAMPLE_TICKETS_B = [
    {
        "summary": "Airflow DAG delayed",
        "ticket_id": "DSSD-29010",
        "jira_url": "https://jira.example.com/browse/DSSD-29010",
        "status": "OPEN Unassigned",
        "slack_text": "slack_link",
        "slack_url": "https://company.slack.com/archives/C456/p444",
    },
    {
        "summary": "CDT batch view negative long runs",
        "ticket_id": "DSSD-29011",
        "jira_url": "https://jira.example.com/browse/DSSD-29011",
        "status": "IN PROGRESS Alice",
        "slack_text": "slack_link",
        "slack_url": "https://company.slack.com/archives/C456/p555",
    },
]

SAMPLE_TTM_TICKETS = [
    {
        "summary": "New issue found during shift",
        "ticket_id": "DRGN-50001",
        "jira_url": "https://jira.example.com/browse/DRGN-50001",
        "status": "OPEN Unassigned",
        "slack_text": "slack_link",
        "slack_url": "https://company.slack.com/archives/C789/p666",
    },
]


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def report_path(tmp_path):
    """Create a two-sheet test report and return its path."""
    wb = Workbook()

    # Night-Shift-NEW (first sheet, rename default)
    ws_night = wb.active
    ws_night.title = "Night-Shift-NEW"
    _build_sheet(
        ws_night,
        sheet_label="Night",
        day=10,
        month="Mar",
        time_range="10AM-10PM",
        from_prev_tickets=SAMPLE_TICKETS_A,
        ttm_tickets=SAMPLE_TTM_TICKETS,
        include_from_prev_header=True,
    )

    # Day-Shift-NEW (second sheet)
    ws_day = wb.create_sheet("Day-Shift-NEW")
    _build_sheet(
        ws_day,
        sheet_label="Day",
        day=10,
        month="Mar",
        time_range="10PM-10AM",
        from_prev_tickets=SAMPLE_TICKETS_B,
        ttm_tickets=[],
        include_from_prev_header=True,
    )

    file_path = tmp_path / "test_report.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def report_no_header(tmp_path):
    """Report where Day-Shift-NEW has no 'from previous shifts' header (v0.1.4 bug)."""
    wb = Workbook()

    ws_night = wb.active
    ws_night.title = "Night-Shift-NEW"
    _build_sheet(
        ws_night,
        sheet_label="Night",
        day=10,
        month="Mar",
        time_range="10AM-10PM",
        from_prev_tickets=SAMPLE_TICKETS_A,
        ttm_tickets=[],
        include_from_prev_header=True,
    )

    ws_day = wb.create_sheet("Day-Shift-NEW")
    _build_sheet(
        ws_day,
        sheet_label="Day",
        day=10,
        month="Mar",
        time_range="10PM-10AM",
        from_prev_tickets=SAMPLE_TICKETS_B,
        ttm_tickets=[],
        include_from_prev_header=False,  # <-- no header
    )

    file_path = tmp_path / "test_report_no_header.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def report_month_boundary(tmp_path):
    """Report at month boundary: Mar 31."""
    wb = Workbook()

    ws_night = wb.active
    ws_night.title = "Night-Shift-NEW"
    _build_sheet(
        ws_night,
        sheet_label="Night",
        day=31,
        month="Mar",
        time_range="10AM-10PM",
        from_prev_tickets=SAMPLE_TICKETS_A[:1],
        ttm_tickets=[],
    )

    ws_day = wb.create_sheet("Day-Shift-NEW")
    _build_sheet(
        ws_day,
        sheet_label="Day",
        day=31,
        month="Mar",
        time_range="10PM-10AM",
        from_prev_tickets=SAMPLE_TICKETS_B[:1],
        ttm_tickets=[],
    )

    file_path = tmp_path / "test_report_month.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def report_dec31(tmp_path):
    """Report at year boundary: Dec 31."""
    wb = Workbook()

    ws_night = wb.active
    ws_night.title = "Night-Shift-NEW"
    _build_sheet(
        ws_night,
        sheet_label="Night",
        day=31,
        month="Dec",
        time_range="10AM-10PM",
        from_prev_tickets=SAMPLE_TICKETS_A[:1],
        ttm_tickets=[],
    )

    ws_day = wb.create_sheet("Day-Shift-NEW")
    _build_sheet(
        ws_day,
        sheet_label="Day",
        day=31,
        month="Dec",
        time_range="10PM-10AM",
        from_prev_tickets=SAMPLE_TICKETS_B[:1],
        ttm_tickets=[],
    )

    file_path = tmp_path / "test_report_dec31.xlsx"
    wb.save(file_path)
    return file_path


# ---------------------------------------------------------------------------
# Jira mock fixture
# ---------------------------------------------------------------------------

@pytest.fixture
def mock_jira(monkeypatch):
    """Mock _jira_get to return controlled responses without network calls."""
    from shift_report import ShiftReport

    _responses = {}

    def set_response(ticket_id: str, status: str, assignee: str, summary: str = ""):
        _responses[ticket_id] = {
            "fields": {
                "status": {"name": status},
                "assignee": {"displayName": assignee} if assignee != "Unassigned" else None,
                "summary": summary or f"Summary for {ticket_id}",
            }
        }

    def mock_get(self, url: str):
        for ticket_id, response in _responses.items():
            if ticket_id in url:
                return response
        return None

    monkeypatch.setattr(ShiftReport, "_jira_get", mock_get)

    return set_response
